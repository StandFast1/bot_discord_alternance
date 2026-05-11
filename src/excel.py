"""Persistent .xlsx file kept in sync with the DB.

Rebuilt after every state change (button) and every scrape cycle. The
file path defaults to /var/lib/alternance-bot/candidatures.xlsx and is
served via /excel slash command.
"""
from __future__ import annotations

import asyncio
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .db import Database


log = logging.getLogger(__name__)


STATE_FILL = {
    "new":      PatternFill("solid", fgColor="D6E4FF"),
    "todo":     PatternFill("solid", fgColor="FFE4B3"),
    "sent":     PatternFill("solid", fgColor="C6F6D5"),
    "rejected": PatternFill("solid", fgColor="FED7D7"),
    "ignored":  PatternFill("solid", fgColor="E2E8F0"),
}

STATE_LABEL_FR = {
    "new": "Nouveau",
    "todo": "À faire",
    "sent": "Envoyée",
    "rejected": "Refusée",
    "ignored": "Ignorée",
}

# (header, db column, width)
COLUMNS = [
    ("ID",                "id",                6),
    ("Source",            "source",           16),
    ("Titre",             "title",            50),
    ("Entreprise",        "company",          28),
    ("Lieu",              "location",         25),
    ("Contrat",           "contract",         16),
    ("État",              "state",            12),
    ("Découverte",        "discovered_at",    20),
    ("Mise à jour état",  "state_changed_at", 20),
    ("URL",               "url",              45),
]


class ExcelExporter:
    def __init__(self, db: Database, path: str):
        self.db = db
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._lock = asyncio.Lock()

    async def rebuild(self) -> None:
        async with self._lock:
            rows = await self.db.all_offers()
            await asyncio.to_thread(self._write_sync, rows)

    def _write_sync(self, rows) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidatures"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2D3748")
        for i, (label, _, _) in enumerate(COLUMNS, start=1):
            c = ws.cell(row=1, column=i, value=label)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        for row_idx, r in enumerate(rows, start=2):
            raw_state = r["state"]
            for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
                val = r[key] if key in r.keys() else None
                if key == "state":
                    val = STATE_LABEL_FR.get(raw_state, raw_state)
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                if key == "state" and raw_state in STATE_FILL:
                    c.fill = STATE_FILL[raw_state]
                if key == "url" and val:
                    c.hyperlink = val
                    c.font = Font(color="0563C1", underline="single")
                if key == "title":
                    c.alignment = Alignment(wrap_text=True, vertical="top")

        for i, (_, _, w) in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[1].height = 24
        ws.freeze_panes = "A2"
        if ws.max_row >= 2:
            ws.auto_filter.ref = ws.dimensions

        tmp = self.path.with_suffix(".xlsx.tmp")
        wb.save(tmp)
        tmp.replace(self.path)
        log.info("xlsx exported: %s (%d rows)", self.path, len(rows))

    async def rebuild_safe(self) -> None:
        """Fire-and-forget rebuild with exception logging."""
        try:
            await self.rebuild()
        except Exception as e:
            log.exception("xlsx rebuild failed: %s", e)
